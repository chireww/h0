"""
AZ 审核结果与金标准对比（Layla 两段式方案） — 自动化测试脚本
============================================================
将 Apifox 测试场景 "AZ_对比审核结果与金标准-调试另一种方案layla" 转换为 Python。

核心设计（区别于其他两个脚本）:
  - 两段式 forEach: 先批量创建所有审核任务收集 taskIds，再逐个获取审核结果
  - 本地读取金标准 Excel（无飞书下载阶段）
  - 审核参数与另外两个版本不同 (material_properties=[2], target_audience=[26])

流程:
  1. 登录 → 提取 token → 解析金标准 Excel → 扫描文件目录
  2. 遍历文件: 上传 → 创建审核任务 → 收集到 taskIds[]
  3. 遍历 taskIds[]: 获取每个任务的审核结果
  4. 与金标准对比 → 输出报告

用法:
  python az_audit_compare_layla.py
"""

import os
import sys
import json
import time
import logging
from pathlib import Path
from typing import Any

import requests
import openpyxl

# ============================================================
# 配置
# ============================================================
CONFIG = {
    "az_base_url": os.getenv("AZ_BASE_URL", "https://dev-api-v3-az-mlr.nullht.com/api"),
    "az_email": os.getenv("AZ_EMAIL", ""),
    "az_password": os.getenv("AZ_PASSWORD", ""),
    "file_dir": os.getenv("AZ_FILE_DIR", r"D:\AZ_test_0"),
    "gold_standard_filepath": os.getenv("AZ_GOLD_STANDARD_PATH", "phase2 金标准测试集.xlsx"),
    "gold_standard_sheetname": os.getenv("AZ_GOLD_STANDARD_SHEET", "金标准集（11+6个审核点）"),
    # Layla 变体特有的审核参数
    "audit_defaults": {
        "literature_package": [],
        "category": "NS",
        "material_properties": [2],
        "file_type_level_1": [73],
        "file_type_level_2": [66],
        "file_type_level_3": [29],
        "medical_education_sub_categories": [4],
        "product_Ids": [
            "f7ce1e4f-5ca7-11f0-8e6d-00163e36469b",
            "f7ce0c74-5ca7-11f0-8e6d-00163e36469b",
        ],
        "target_audience": [26],
    },
    "poll_interval": 3,
    "max_poll_attempts": 120,
    "verify_ssl": False,
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("az_layla")


# ============================================================
# AZ 客户端
# ============================================================
class AZClient:
    def __init__(self, config: dict):
        self.base_url = config["az_base_url"].rstrip("/")
        self.verify_ssl = config["verify_ssl"]
        self.audit_defaults = config["audit_defaults"]
        self.session = requests.Session()
        self.session.verify = self.verify_ssl
        self.session.headers.update({
            "User-Agent": "AZ-Audit-Layla/1.0 (Python)",
            "Content-Type": "application/json",
        })
        self.token: str | None = None

    @property
    def auth_header(self) -> dict:
        return {"Authorization": self.token} if self.token else {}

    # ----- 登录 -----
    def login(self, email: str, password: str) -> str:
        url = f"{self.base_url}/auth/email/login"
        body = {"email": email, "password": password}
        log.info(f"🔑 登录: {email}")
        resp = self.session.post(url, json=body)
        resp.raise_for_status()
        data = resp.json()
        token = data.get("data", {}).get("token")
        if not token:
            raise RuntimeError(f"登录失败: {resp.text[:300]}")
        self.token = f"Bearer {token}"
        log.info("✅ 登录成功")
        return self.token

    # ----- 上传文件 -----
    def upload_file(self, file_path: str) -> str | None:
        url = f"{self.base_url}/oss/upload"
        file_name = Path(file_path).name
        log.info(f"📤 上传: {file_name}")
        try:
            with open(file_path, "rb") as f:
                resp = self.session.post(
                    url, headers=self.auth_header,
                    files={"file": (file_name, f)},
                )
            resp.raise_for_status()
            file_id = resp.json().get("data")
            if file_id:
                log.info(f"   ✅ file_id={file_id}")
                return file_id
            log.warning(f"   ⚠️ 空响应: {resp.text[:200]}")
            return None
        except requests.RequestException as e:
            log.error(f"   ❌ 上传失败: {e}")
            return None

    # ----- 创建审核任务 -----
    def create_audit_task(self, file_id: str, file_name: str) -> str | None:
        """创建审核任务，返回 taskId。"""
        url = f"{self.base_url}/audit/management/add"
        body = {"file_id": file_id, "file_name": file_name, **self.audit_defaults}
        log.info(f"📝 创建审核: {file_name}")
        try:
            resp = self.session.post(url, json=body, headers=self.auth_header)
            resp.raise_for_status()
            task_id = resp.json().get("data")
            if task_id:
                log.info(f"   ✅ taskId={task_id}")
                return task_id
            log.warning(f"   ⚠️ 空响应: {resp.text[:200]}")
            return None
        except requests.RequestException as e:
            log.error(f"   ❌ 创建失败: {e}")
            return None

    # ----- 获取审核结果 -----
    def get_audit_detail(self, task_id: str) -> dict | None:
        url = f"{self.base_url}/audit/management/detail"
        body = {"id": task_id}
        try:
            resp = self.session.post(url, json=body, headers=self.auth_header)
            resp.raise_for_status()
            return resp.json()
        except requests.RequestException as e:
            log.error(f"   ❌ 获取失败 (id={task_id}): {e}")
            return None


# ============================================================
# 金标准解析
# ============================================================
def read_excel_gold_standard(filepath: str, sheet_name: str) -> tuple[dict[str, list[dict]], int]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}")
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append({headers[i]: value for i, value in enumerate(row) if i < len(headers) and headers[i] is not None})
    wb.close()

    merged: dict[str, list[dict]] = {}
    for row in rows:
        fn = row.get("文件名") or row.get("文件名称") or row.get("file_name")
        if fn:
            merged.setdefault(fn, []).append(row)
    log.info(f"✅ 金标准: {len(rows)} 条记录, {len(merged)} 个文件")
    return merged, len(merged)


def scan_files(directory: str) -> list[str]:
    p = Path(directory)
    if not p.exists():
        raise FileNotFoundError(f"目录不存在: {directory}")
    files = [str(fp.resolve()) for fp in p.iterdir() if fp.is_file() and fp.name != ".DS_Store"]
    log.info(f"📂 扫描到 {len(files)} 个文件")
    return files


# ============================================================
# 对比逻辑
# ============================================================
def compare_with_gold_standard(
    client: AZClient,
    task_ids: list[str],
    file_names: dict[str, str],  # task_id → file_name 映射
    gold_standard: dict[str, list[dict]],
    poll_interval: int = 3,
    max_attempts: int = 120,
) -> list[dict]:
    """
    Layla 方案核心：逐个 taskId 获取审核结果，与金标准对比。

    这是原 Apifox 脚本未实现的部分，此处补齐。
    """
    results = []
    for task_id in task_ids:
        file_name = file_names.get(task_id, task_id)
        log.info(f"🔍 获取审核结果: {file_name} (taskId={task_id})")

        # 轮询等待审核完成
        detail = None
        for attempt in range(max_attempts):
            detail = client.get_audit_detail(task_id)
            if detail is None:
                break
            status = (
                detail.get("data", {}).get("status")
                or detail.get("data", {}).get("audit_status")
                or detail.get("data", {}).get("state", "")
            )
            if status in ("completed", "done", "finished", "审核完成", "已完成"):
                break
            if attempt == 0:
                log.info(f"   ⏳ 等待审核完成...")
            time.sleep(poll_interval)
        else:
            log.warning(f"   ⚠️ 轮询超时，使用最后一次结果")

        if detail is None:
            results.append({"file_name": file_name, "task_id": task_id, "status": "error", "error": "无法获取审核结果"})
            continue

        audit_data = detail.get("data", {})
        gold_rows = gold_standard.get(file_name) or gold_standard.get(Path(file_name).stem, [])

        if not gold_rows:
            log.warning(f"   ⚠️ 无金标准数据")
            results.append({"file_name": file_name, "task_id": task_id, "status": "no_gold_standard", "audit_data": audit_data})
            continue

        row_results = []
        for gold_row in gold_rows:
            point = gold_row.get("审核点", gold_row.get("审核项", ""))
            expected = gold_row.get("期望值", gold_row.get("期望结果", gold_row.get("金标准", "")))
            actual = _extract_value(audit_data, gold_row)
            passed = str(actual) == str(expected) if actual is not None else False
            row_results.append({"审核点": point, "期望值": expected, "实际值": actual, "通过": passed})

        passed = sum(1 for r in row_results if r["通过"])
        total = len(row_results)
        log.info(f"   {'✅' if passed == total else '❌'} {passed}/{total}")
        results.append({
            "file_name": file_name, "task_id": task_id, "status": "completed",
            "points": row_results, "passed": passed, "total": total, "pass_rate": f"{passed}/{total}",
        })
    return results


def _extract_value(audit_data: dict, gold_row: dict) -> Any | None:
    field = gold_row.get("字段名") or gold_row.get("field") or gold_row.get("审核字段")
    if field:
        value = audit_data
        for key in str(field).split("."):
            if isinstance(value, dict):
                value = value.get(key)
            else:
                return None
        return value
    return _deep_search(audit_data, gold_row.get("审核点", ""))


def _deep_search(data: dict, target: str) -> Any | None:
    if not isinstance(data, dict):
        return None
    for key, value in data.items():
        if target in str(key):
            return value
        if isinstance(value, dict):
            r = _deep_search(value, target)
            if r is not None:
                return r
    return None


def generate_report(results: list[dict], output: str = "audit_comparison_report_layla.json"):
    total = len(results)
    completed = sum(1 for r in results if r["status"] == "completed")
    errors = sum(1 for r in results if r["status"] == "error")
    no_gold = sum(1 for r in results if r["status"] == "no_gold_standard")
    all_passed = sum(r.get("passed", 0) for r in results if r["status"] == "completed")
    all_total = sum(r.get("total", 0) for r in results if r["status"] == "completed")

    summary = {
        "方案": "Layla 两段式",
        "总文件数": total,
        "成功对比": completed,
        "获取失败": errors,
        "无金标准": no_gold,
        "审核点通过率": f"{all_passed}/{all_total}" if all_total else "N/A",
    }
    report = {"summary": summary, "details": results}
    with open(output, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    log.info("=" * 60)
    log.info("📊 对比报告 (Layla)")
    for k, v in summary.items():
        log.info(f"  {k}: {v}")
    log.info(f"  详细结果: {output}")
    return report


# ============================================================
# 主流程
# ============================================================
def main():
    cfg = CONFIG
    if not cfg["az_email"] or not cfg["az_password"]:
        log.error("❌ 请设置 AZ_EMAIL 和 AZ_PASSWORD 环境变量")
        sys.exit(1)

    az = AZClient(cfg)

    # ===== Step 1: 登录 + 加载金标准 + 扫描文件 =====
    log.info("=" * 60)
    log.info("Step 1: 登录 & 加载数据")
    log.info("=" * 60)

    # 1a. 先解析金标准 Excel（原脚本在 login pre-request 中做）
    gold_standard, file_count = read_excel_gold_standard(
        cfg["gold_standard_filepath"], cfg["gold_standard_sheetname"]
    )

    # 1b. 登录
    az.login(cfg["az_email"], cfg["az_password"])

    # 1c. 扫描文件目录（原脚本在 login post-request 中做）
    file_paths = scan_files(cfg["file_dir"])
    if not file_paths:
        log.error("❌ 没有待审核文件")
        sys.exit(1)

    # ===== Step 2: 第一段循环 — 上传 + 创建审核任务 → 收集 taskIds =====
    log.info("=" * 60)
    log.info("Step 2: 上传文件 & 创建审核任务 → 收集 taskIds")
    log.info("=" * 60)

    task_ids: list[str] = []          # 原脚本的 taskIds 数组
    file_names: dict[str, str] = {}   # task_id → file_name 映射（方便后续对比）

    for fp in file_paths:
        fname = Path(fp).name

        # 上传
        file_id = az.upload_file(fp)
        if not file_id:
            log.warning(f"⏭️ 跳过 {fname}（上传失败）")
            continue

        # 创建审核任务（对应原脚本 IF upload成功 → 创建任务）
        task_id = az.create_audit_task(file_id, fname)
        if task_id:
            task_ids.append(task_id)           # push 到 taskIds
            file_names[task_id] = fname
        else:
            log.warning(f"⏭️ {fname} 上传成功但创建任务失败")

    log.info(f"📋 收集到 {len(task_ids)} 个 taskId")

    if not task_ids:
        log.error("❌ 没有成功创建任何审核任务")
        sys.exit(1)

    # ===== Step 3: 第二段循环 — 逐个获取审核结果 + 对比金标准 =====
    log.info("=" * 60)
    log.info("Step 3: 遍历 taskIds → 获取审核结果 & 对比金标准")
    log.info("=" * 60)
    log.info("   (原 Apifox 脚本中此步骤仅为裸 API 调用，此处补齐对比逻辑)")

    results = compare_with_gold_standard(
        az, task_ids, file_names, gold_standard,
        poll_interval=cfg["poll_interval"],
        max_attempts=cfg["max_poll_attempts"],
    )

    # ===== Step 4: 报告 =====
    generate_report(results)

    all_pass = all(
        r["status"] == "completed" and r.get("passed", 0) == r.get("total", 0)
        for r in results
    )
    sys.exit(0 if all_pass else 1)


if __name__ == "__main__":
    main()
