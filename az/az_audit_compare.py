"""
AZ 审核结果与金标准对比 — 自动化测试脚本
============================================
将 Apifox 测试场景转换为 Python 实现。

流程:
  1. 登录获取 token
  2. 解析 Excel 金标准数据
  3. 扫描待审核文件目录
  4. 遍历: 上传文件 → 创建审核任务 → 收集 task_id
  5. 获取所有任务的审核结果
  6. 与金标准对比，输出报告

用法:
  python az_audit_compare.py

配置通过环境变量或下方 CONFIG 字典设置。
"""

import os
import sys
import json
import time
import logging
from pathlib import Path
from typing import Any

import requests
import openpyxl

# ============================================================
# 配置 — 按实际环境修改
# ============================================================
CONFIG = {
    "base_url": os.getenv("AZ_BASE_URL", "https://dev-api-v3-az-mlr.nullht.com/api"),
    "email": os.getenv("AZ_EMAIL", ""),
    "password": os.getenv("AZ_PASSWORD", ""),
    # 待审核文件所在目录
    "file_dir": os.getenv("AZ_FILE_DIR", r"D:\AZ_test_0"),
    # 金标准 Excel 路径
    "gold_standard_filepath": os.getenv("AZ_GOLD_STANDARD_PATH", "phase2 金标准测试集.xlsx"),
    "gold_standard_sheetname": os.getenv("AZ_GOLD_STANDARD_SHEET", "金标准集（11+6个审核点）"),
    # 审核请求体中的固定参数（来自原 Apifox 脚本）
    "audit_defaults": {
        "literature_package": [],
        "category": "NS",
        "material_properties": [1515],
        "file_type_level_1": [89],
        "file_type_level_2": [80],
        "file_type_level_3": [35],
        "medical_education_sub_categories": [1517],
        "product_Ids": [
            "f7ce1e4f-5ca7-11f0-8e6d-00163e36469b",
            "f7ce0c74-5ca7-11f0-8e6d-00163e36469b",
        ],
        "target_audience": [1534],
    },
    # 轮询审核结果的间隔（秒）
    "poll_interval": 3,
    # 最大轮询次数
    "max_poll_attempts": 60,
    # 是否跳过 SSL 证书验证
    "verify_ssl": False,
}

# ============================================================
# 日志
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("az_audit")


# ============================================================
# 工具函数
# ============================================================
def read_excel_gold_standard(
    filepath: str, sheet_name: str
) -> tuple[dict[str, Any], int]:
    """
    读取金标准 Excel 文件，返回 (合并后的金标准数据, 文件数量)。

    金标准 Excel 的结构:
      - 每行代表一个审核点
      - 包含文件名、期望的审核结果等字段
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' 不存在，可用的: {wb.sheetnames}")

    ws = wb[sheet_name]

    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # 读取所有数据行
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):  # 跳过全空行
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i] is not None:
                    row_dict[headers[i]] = value
            rows.append(row_dict)

    wb.close()

    # 按文件名分组，构造 merged_result
    merged_result = {}
    for row in rows:
        file_name = row.get("文件名") or row.get("文件名称") or row.get("file_name")
        if file_name is None:
            continue
        if file_name not in merged_result:
            merged_result[file_name] = []
        merged_result[file_name].append(row)

    log.info(f"✅ 加载金标准: {len(rows)} 条记录, {len(merged_result)} 个文件")
    return merged_result, len(merged_result)


def scan_files(directory: str) -> list[str]:
    """扫描目录下所有文件（排除 .DS_Store），返回完整路径列表。"""
    dir_path = Path(directory)
    if not dir_path.exists():
        raise FileNotFoundError(f"目录不存在: {directory}")

    files = [
        str(p.resolve())
        for p in dir_path.iterdir()
        if p.is_file() and p.name != ".DS_Store"
    ]
    log.info(f"📂 扫描到 {len(files)} 个文件")
    return files


def extract_filename(file_path: str) -> str:
    """从完整路径中提取文件名。"""
    return Path(file_path).name


# ============================================================
# API 客户端
# ============================================================
class AZClient:
    """AZ 审核系统 API 客户端"""

    def __init__(self, config: dict):
        self.base_url = config["base_url"].rstrip("/")
        self.verify_ssl = config["verify_ssl"]
        self.session = requests.Session()
        self.session.verify = self.verify_ssl
        self.session.headers.update(
            {
                "User-Agent": "AZ-Audit-Test/1.0 (Python)",
                "Content-Type": "application/json",
            }
        )
        self.token: str | None = None
        self.config = config

    def _auth_header(self) -> dict:
        return {"Authorization": self.token} if self.token else {}

    # ---------- 1. 登录 ----------
    def login(self, email: str, password: str) -> str:
        """登录并返回 token"""
        url = f"{self.base_url}/auth/email/login"
        body = {"email": email, "password": password}

        log.info(f"🔑 登录中... (email={email})")
        resp = self.session.post(url, json=body)
        resp.raise_for_status()

        data = resp.json()
        token = data.get("data", {}).get("token")
        if not token:
            raise RuntimeError(f"登录失败: 响应中无 token — {resp.text[:500]}")

        self.token = f"Bearer {token}"
        log.info("✅ 登录成功")
        return self.token

    # ---------- 2. 上传文件 ----------
    def upload_file(self, file_path: str) -> str | None:
        """上传文件，返回 file_id（失败返回 None）。"""
        url = f"{self.base_url}/oss/upload"
        headers = self._auth_header()
        # 上传文件时不设置 Content-Type，让 requests 自动处理 multipart
        del headers["Content-Type"]

        file_name = extract_filename(file_path)
        log.info(f"📤 上传: {file_name}")

        try:
            with open(file_path, "rb") as f:
                resp = self.session.post(
                    url,
                    headers=headers,
                    files={"file": (file_name, f)},
                )
            resp.raise_for_status()
            file_id = resp.json().get("data")
            if file_id:
                log.info(f"   ✅ file_id={file_id}")
                return file_id
            else:
                log.warning(f"   ⚠️ 上传返回空 data: {resp.text[:200]}")
                return None
        except requests.RequestException as e:
            log.error(f"   ❌ 上传失败: {e}")
            return None
        finally:
            # 恢复 Content-Type
            self.session.headers["Content-Type"] = "application/json"

    # ---------- 3. 创建审核任务 ----------
    def create_audit_task(self, file_id: str, file_name: str) -> str | None:
        """创建审核任务，返回 task_id（失败返回 None）。"""
        url = f"{self.base_url}/audit/management/add"

        body = {
            "file_id": file_id,
            "file_name": file_name,
            **self.config["audit_defaults"],
        }

        log.info(f"📝 创建审核任务: {file_name}")
        try:
            resp = self.session.post(url, json=body)
            resp.raise_for_status()
            task_id = resp.json().get("data")
            if task_id:
                log.info(f"   ✅ task_id={task_id}")
                return task_id
            else:
                log.warning(f"   ⚠️ 创建任务返回空 data: {resp.text[:200]}")
                return None
        except requests.RequestException as e:
            log.error(f"   ❌ 创建任务失败: {e}")
            return None

    # ---------- 4. 获取审核结果详情 ----------
    def get_audit_detail(self, task_id: str) -> dict | None:
        """获取单个审核任务的详情（含审核结果）。"""
        url = f"{self.base_url}/audit/management/detail"
        body = {"id": task_id}

        try:
            resp = self.session.post(url, json=body)
            resp.raise_for_status()
            return resp.json()
        except requests.RequestException as e:
            log.error(f"   ❌ 获取审核详情失败 (id={task_id}): {e}")
            return None


# ============================================================
# 对比逻辑
# ============================================================
def compare_with_gold_standard(
    client: AZClient,
    task_queue: list[dict],
    gold_standard: dict[str, list[dict]],
) -> list[dict]:
    """
    获取所有任务的审核结果，与金标准对比。

    task_queue: [{"id": task_id, "fileName": name}, ...]
    gold_standard: {"文件名": [{"审核点": ..., "期望值": ..., ...}, ...], ...}

    返回对比结果列表。
    """
    comparison_results = []

    for task in task_queue:
        task_id = task["id"]
        file_name = task["fileName"]
        log.info(f"🔍 获取审核结果: {file_name} (id={task_id})")

        detail = client.get_audit_detail(task_id)
        if detail is None:
            comparison_results.append(
                {
                    "file_name": file_name,
                    "task_id": task_id,
                    "status": "error",
                    "error": "无法获取审核结果",
                }
            )
            continue

        # 从响应中提取审核结果数据
        audit_data = detail.get("data", {})

        # 获取该文件对应的金标准
        gold_rows = gold_standard.get(file_name, [])
        if not gold_rows:
            # 尝试模糊匹配：去掉扩展名
            stem = Path(file_name).stem
            gold_rows = gold_standard.get(stem, [])

        if not gold_rows:
            log.warning(f"   ⚠️ 未找到 '{file_name}' 的金标准数据")
            comparison_results.append(
                {
                    "file_name": file_name,
                    "task_id": task_id,
                    "status": "no_gold_standard",
                    "audit_data": audit_data,
                }
            )
            continue

        # 逐项对比每个审核点
        row_results = []
        for gold_row in gold_rows:
            point_name = gold_row.get("审核点", gold_row.get("审核项", ""))
            expected = gold_row.get("期望值", gold_row.get("期望结果", gold_row.get("金标准", "")))
            # 从审核结果中提取对应字段的值
            # （字段映射需要根据实际 API 响应结构调整）
            actual = _extract_audit_value(audit_data, gold_row)

            passed = str(actual) == str(expected) if actual is not None else False
            row_results.append(
                {
                    "审核点": point_name,
                    "期望值": expected,
                    "实际值": actual,
                    "通过": passed,
                }
            )

        passed_count = sum(1 for r in row_results if r["通过"])
        total_count = len(row_results)
        log.info(
            f"   {'✅' if passed_count == total_count else '❌'} "
            f"{passed_count}/{total_count} 通过"
        )

        comparison_results.append(
            {
                "file_name": file_name,
                "task_id": task_id,
                "status": "completed",
                "points": row_results,
                "passed": passed_count,
                "total": total_count,
                "pass_rate": f"{passed_count}/{total_count}",
            }
        )

    return comparison_results


def _extract_audit_value(audit_data: dict, gold_row: dict) -> Any | None:
    """
    从审核结果数据中提取对应审核点的实际值。

    这里需要根据实际的 API 响应数据结构来映射。
    常见情况：
      - gold_row 中有一个"字段名"列，指定 API 响应中的 JSONPath
      - 或者根据"审核点"名称在 audit_data 中查找匹配的 key
    """
    # 尝试从 gold_row 中获取字段映射
    field_path = gold_row.get("字段名") or gold_row.get("field") or gold_row.get("审核字段")

    if field_path:
        # 支持点号分隔的路径，如 "data.audit_result.category"
        value = audit_data
        for key in str(field_path).split("."):
            if isinstance(value, dict):
                value = value.get(key)
            else:
                return None
        return value

    # 回退：尝试用审核点名称在 audit_data 中匹配
    point_name = gold_row.get("审核点", "")
    if isinstance(audit_data, dict):
        # 递归搜索
        result = _deep_search(audit_data, point_name)
        if result is not None:
            return result

    return None


def _deep_search(data: dict, target_key: str) -> Any | None:
    """在嵌套字典中递归搜索匹配的 key（模糊匹配）。"""
    if not isinstance(data, dict):
        return None
    for key, value in data.items():
        if target_key in str(key):
            return value
        if isinstance(value, dict):
            result = _deep_search(value, target_key)
            if result is not None:
                return result
    return None


# ============================================================
# 报告生成
# ============================================================
def generate_report(results: list[dict], output_path: str = "audit_comparison_report.json"):
    """生成对比报告（JSON）并输出摘要。"""
    total = len(results)
    completed = sum(1 for r in results if r["status"] == "completed")
    errors = sum(1 for r in results if r["status"] == "error")
    no_gold = sum(1 for r in results if r["status"] == "no_gold_standard")

    all_passed = 0
    all_total = 0
    for r in results:
        if r["status"] == "completed":
            all_passed += r.get("passed", 0)
            all_total += r.get("total", 0)

    summary = {
        "总文件数": total,
        "成功对比": completed,
        "获取失败": errors,
        "无金标准": no_gold,
        "审核点通过率": f"{all_passed}/{all_total}" if all_total > 0 else "N/A",
    }

    report = {
        "summary": summary,
        "details": results,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    log.info("=" * 60)
    log.info("📊 对比报告摘要")
    log.info("=" * 60)
    for k, v in summary.items():
        log.info(f"  {k}: {v}")
    log.info(f"  详细结果已保存至: {output_path}")
    return report


# ============================================================
# 主流程
# ============================================================
def main():
    config = CONFIG

    # 基础校验
    if not config["email"] or not config["password"]:
        log.error("❌ 请设置 AZ_EMAIL 和 AZ_PASSWORD 环境变量，或直接修改 CONFIG")
        sys.exit(1)

    client = AZClient(config)

    # ===== Step 1: 读取金标准 Excel =====
    log.info("=" * 60)
    log.info("Step 1: 加载金标准数据")
    log.info("=" * 60)
    gold_standard, gold_file_count = read_excel_gold_standard(
        config["gold_standard_filepath"], config["gold_standard_sheetname"]
    )

    # ===== Step 2: 登录 =====
    log.info("=" * 60)
    log.info("Step 2: 登录")
    log.info("=" * 60)
    client.login(config["email"], config["password"])

    # ===== Step 3: 扫描待审核文件 =====
    log.info("=" * 60)
    log.info("Step 3: 扫描待审核文件")
    log.info("=" * 60)
    file_paths = scan_files(config["file_dir"])
    if not file_paths:
        log.error("❌ 没有找到待审核文件")
        sys.exit(1)

    # ===== Step 4: 遍历 — 上传 + 创建审核任务 =====
    log.info("=" * 60)
    log.info("Step 4: 上传文件 & 创建审核任务")
    log.info("=" * 60)
    task_queue: list[dict] = []

    for file_path in file_paths:
        file_name = extract_filename(file_path)

        # 上传
        file_id = client.upload_file(file_path)
        if not file_id:
            log.warning(f"⏭️ 跳过 {file_name}（上传失败）")
            continue

        # 创建审核任务
        task_id = client.create_audit_task(file_id, file_name)
        if task_id:
            task_queue.append({"id": task_id, "fileName": file_name})
        else:
            log.warning(f"⏭️ {file_name} 上传成功但创建任务失败")

    log.info(f"📋 共创建 {len(task_queue)} 个审核任务")

    # ===== Step 5: 等待审核完成（轮询） =====
    log.info("=" * 60)
    log.info("Step 5: 等待审核完成")
    log.info("=" * 60)

    pending_ids = set(task["id"] for task in task_queue)
    max_attempts = config["max_poll_attempts"]
    interval = config["poll_interval"]

    for attempt in range(max_attempts):
        if not pending_ids:
            break

        still_pending = set()
        for task_id in pending_ids:
            detail = client.get_audit_detail(task_id)
            if detail is None:
                still_pending.add(task_id)
                continue

            # 检查审核状态（根据实际 API 响应结构调整）
            data = detail.get("data", {})
            status = data.get("status") or data.get("audit_status") or data.get("state")
            if status in ("completed", "done", "finished", "审核完成", "已完成"):
                log.info(f"   ✅ 任务 {task_id} 审核完成")
            else:
                still_pending.add(task_id)

        pending_ids = still_pending

        if pending_ids:
            log.info(
                f"   ⏳ 等待中... ({len(pending_ids)} 个任务未完成, "
                f"第 {attempt + 1}/{max_attempts} 次检查)"
            )
            time.sleep(interval)

    if pending_ids:
        log.warning(f"⚠️ {len(pending_ids)} 个任务在超时后仍未完成，将继续获取结果")

    # ===== Step 6: 获取审核结果并对比 =====
    log.info("=" * 60)
    log.info("Step 6: 获取审核结果 & 对比金标准")
    log.info("=" * 60)
    comparison_results = compare_with_gold_standard(client, task_queue, gold_standard)

    # ===== Step 7: 生成报告 =====
    log.info("=" * 60)
    log.info("Step 7: 生成报告")
    log.info("=" * 60)
    generate_report(comparison_results)

    # 返回退出码
    all_pass = all(
        r["status"] == "completed" and r.get("passed", 0) == r.get("total", 0)
        for r in comparison_results
    )
    sys.exit(0 if all_pass else 1)


if __name__ == "__main__":
    main()
