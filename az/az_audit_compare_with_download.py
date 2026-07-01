"""
AZ 审核结果与金标准对比（含飞书文件夹下载） — 自动化测试脚本
============================================================
将 Apifox 测试场景 "AZ_对比审核结果与金标准(添加下载文件夹)" 转换为 Python。

完整流程:
  Phase 1 — 从飞书下载金标准 Excel:
    1. 创建文件夹下载任务    POST /space/api/box/invoke/create/
    2. 轮询下载状态           GET  /space/api/box/invoke/check/  → invoke_status == 1
    3. 打包文件夹为 zip       POST /space/api/box/zip/create/
    4. 轮询 zip 状态          GET  /space/api/box/zip/check/     → zip_status == 0
    5. 下载并解压 zip，获取 Excel token
    6. 导出 Excel 文件        POST /space/api/export/create/
    7. 轮询导出结果           GET  /space/api/export/result/{ticket} → job_error_msg == "success"
    8. 下载 Excel 并解析金标准数据

  Phase 2 — 审核流程:
    9. 登录 AZ 系统          POST /api/auth/email/login
   10. 扫描待审核文件目录
   11. 遍历: 上传文件 → 创建审核任务
   12. 获取所有审核结果 → 与金标准对比 → 输出报告

用法:
  python az_audit_compare_with_download.py

配置通过环境变量或下方 CONFIG 字典设置。
"""

import os
import sys
import json
import time
import logging
import zipfile
import tempfile
import io
from pathlib import Path
from typing import Any

import requests
import openpyxl

# ============================================================
# 配置
# ============================================================
CONFIG = {
    # ---- 飞书 / Lark ----
    "feishu_host": os.getenv("FEISHU_HOST", "https://wuk7uud7j2.feishu.cn"),
    "feishu_cookies": os.getenv("FEISHU_COOKIES", ""),
    "feishu_referer": os.getenv("FEISHU_REFERER", ""),
    "feishu_csrf_token": os.getenv("FEISHU_CSRF_TOKEN", ""),
    "feishu_folder_name": os.getenv("FEISHU_FOLDER_NAME", "赵庆贺"),
    "feishu_folder_token": os.getenv("FEISHU_FOLDER_TOKEN", "Q2aefwtVtlQGxndzioGc2k83nTc"),
    "feishu_file_token": os.getenv("FEISHU_FILE_TOKEN", "RAy8sfHfGhPdrttNJX4cLh7lnWf"),
    "feishu_excel_token": os.getenv("FEISHU_EXCEL_TOKEN", "NmFcsY3XYhNLXJtPikrc9eUWnJh"),
    "feishu_download_path": os.getenv("FEISHU_DOWNLOAD_PATH", ""),
    # ---- AZ 审核系统 ----
    "az_base_url": os.getenv("AZ_BASE_URL", "https://dev-api-v3-az-mlr.nullht.com/api"),
    "az_email": os.getenv("AZ_EMAIL", ""),
    "az_password": os.getenv("AZ_PASSWORD", ""),
    "file_dir": os.getenv("AZ_FILE_DIR", r"D:\AZ_test_0"),
    # ---- 金标准 ----
    "gold_standard_sheetname": os.getenv("AZ_GOLD_STANDARD_SHEET", "金标准集（11+6个审核点）"),
    "gold_standard_filename": os.getenv("AZ_GOLD_STANDARD_FILENAME", "phase2 金标准测试集.xlsx"),
    # ---- 审核请求默认参数（来自原始脚本） ----
    "audit_defaults": {
        "literature_package": [],
        "category": "NS",
        "material_properties": [2],
        "file_type_level_1": [73],
        "file_type_level_2": [66],
        "file_type_level_3": [29],
        "medical_education_sub_categories": [4],
        "product_Ids": [
            "f7ce1e4f-5ca7-11f0-8e6d-00163e36469b",
            "f7ce0c74-5ca7-11f0-8e6d-00163e36469b",
        ],
        "target_audience": [1534],
    },
    # ---- 轮询参数 ----
    "poll_interval": 3,
    "max_poll_attempts": 120,
    "verify_ssl": False,
}

# ============================================================
# 日志
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("az_audit")


# ============================================================
# Phase 1: 飞书 API 客户端 — 文件夹下载 → Excel
# ============================================================
class FeishuClient:
    """飞书/Lark 文件夹操作客户端"""

    def __init__(self, config: dict):
        self.host = config["feishu_host"].rstrip("/")
        self.cookies = config["feishu_cookies"]
        self.referer = config["feishu_referer"]
        self.csrf_token = config["feishu_csrf_token"]
        self.verify_ssl = config["verify_ssl"]
        self.folder_name = config["feishu_folder_name"]
        self.folder_token = config["feishu_folder_token"]
        self.excel_token = config["feishu_excel_token"]
        self.download_path = config["feishu_download_path"]
        self.sheet_name = config["gold_standard_sheetname"]

        self.session = requests.Session()
        self.session.verify = self.verify_ssl
        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Cookie": self.cookies,
                "Referer": self.referer or self.host,
                "X-CSRF-TOKEN": self.csrf_token,
            }
        )

    # ----- Step 1: 创建文件夹下载任务 -----
    def create_folder_download(self) -> str:
        """
        POST /space/api/box/invoke/create/
        返回 invoke_code
        """
        url = f"{self.host}/space/api/box/invoke/create/"
        form_data = {
            "invoke_data": json.dumps(
                [
                    {
                        "token": self.folder_name,
                        "type": 0,
                        "name": self.folder_name,
                        "ownerType": 5,
                    }
                ]
            )
        }
        log.info(f"📁 创建文件夹下载任务: {self.folder_name}")
        resp = self.session.post(url, data=form_data)
        resp.raise_for_status()
        result = resp.json()
        invoke_code = result.get("data", {}).get("invoke_code")
        if not invoke_code:
            raise RuntimeError(f"创建文件夹下载失败: {resp.text[:300]}")
        log.info(f"   ✅ invoke_code={invoke_code}")
        return invoke_code

    # ----- Step 2: 轮询文件夹下载状态 -----
    def poll_folder_download(self, invoke_code: str, timeout: int = 600) -> int:
        """
        GET /space/api/box/invoke/check/?invoke_code=...
        返回 invoke_status (1 = 完成)
        """
        url = f"{self.host}/space/api/box/invoke/check/"
        params = {"invoke_code": invoke_code}

        start = time.time()
        while time.time() - start < timeout:
            resp = self.session.get(url, params=params)
            resp.raise_for_status()
            result = resp.json()
            status = result.get("data", {}).get("invoke_status", -1)
            log.info(f"   ⏳ 文件夹下载状态: invoke_status={status}")
            if status == 1:
                log.info("   ✅ 文件夹下载完成")
                return status
            time.sleep(CONFIG["poll_interval"])

        raise TimeoutError(f"文件夹下载超时 ({timeout}s)")

    # ----- Step 3: 创建文件夹 zip -----
    def create_zip(self) -> str:
        """
        POST /space/api/box/zip/create/
        返回 zip_code
        """
        url = f"{self.host}/space/api/box/zip/create/"
        body = [
            {
                "obj_token": self.folder_token,
                "obj_type": 0,
                "owner_type": 5,
            }
        ]
        log.info(f"🗜️ 创建 zip 包: folder_token={self.folder_token}")
        resp = self.session.post(url, json=body)
        resp.raise_for_status()
        result = resp.json()
        zip_code = result.get("data", {}).get("zip_code")
        if not zip_code:
            raise RuntimeError(f"创建 zip 失败: {resp.text[:300]}")
        log.info(f"   ✅ zip_code={zip_code}")
        return zip_code

    # ----- Step 4: 轮询 zip 状态 -----
    def poll_zip_status(self, zip_code: str, timeout: int = 600) -> int:
        """
        GET /space/api/box/zip/check/?zip_code=...
        返回 zip_status (0 = 完成)
        """
        url = f"{self.host}/space/api/box/zip/check/"
        params = {"zip_code": zip_code}

        start = time.time()
        while time.time() - start < timeout:
            resp = self.session.get(url, params=params)
            resp.raise_for_status()
            result = resp.json()
            status = result.get("data", {}).get("zip_status", -1)
            log.info(f"   ⏳ zip 状态: zip_status={status}")
            if status == 0:
                log.info("   ✅ zip 打包完成")
                return status
            time.sleep(CONFIG["poll_interval"])

        raise TimeoutError(f"zip 打包超时 ({timeout}s)")

    # ----- Step 5: 下载 zip 并提取 Excel token -----
    def download_zip_and_get_excel_token(self, zip_code: str) -> str:
        """
        下载 zip 文件，解压后在文件夹中找到 Excel 文件并获取其 token。

        原 Apifox 调用 extract_zip_and_get_excel_url.py 实现，
        此处内联实现。
        """
        # 飞书 zip 下载 URL 模式
        # 实际 URL 需要根据飞书 API 响应构造，这里假设 zip_code 可直接用于下载
        zip_url = f"{self.host}/space/api/box/zip/download/"
        params = {"zip_code": zip_code}

        log.info(f"📥 下载 zip 并提取 Excel token...")

        # 下载 zip
        resp = self.session.get(zip_url, params=params)
        resp.raise_for_status()

        # 保存到临时目录
        download_dir = Path(self.download_path) if self.download_path else Path(tempfile.gettempdir()) / "feishu_download"
        download_dir.mkdir(parents=True, exist_ok=True)

        zip_path = download_dir / f"{self.folder_name}.zip"
        with open(zip_path, "wb") as f:
            f.write(resp.content)
        log.info(f"   zip 已保存: {zip_path} ({len(resp.content)} bytes)")

        # 解压
        extract_dir = download_dir / self.folder_name
        extract_dir.mkdir(exist_ok=True)
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(extract_dir)
        log.info(f"   解压至: {extract_dir}")

        # 查找 Excel 文件
        excel_files = list(extract_dir.glob("**/*.xlsx")) + list(extract_dir.glob("**/*.xls"))
        if not excel_files:
            raise FileNotFoundError(f"在 {extract_dir} 中未找到 Excel 文件")

        excel_path = excel_files[0]
        log.info(f"   找到 Excel: {excel_path.name}")

        # 如果解压出来的文件名与 folder_name 匹配，
        # Excel token 就是在飞书中的文件 token
        # （这里假设 folder_token 对应的就是 Excel 文件，实际逻辑取决于 API 响应）
        excel_token = self.excel_token  # 使用配置中的 token
        log.info(f"   ✅ Excel token: {excel_token}")
        return excel_token

    # ----- Step 6: 创建文件导出任务 -----
    def create_export(self, excel_token: str) -> str:
        """
        POST /space/api/export/create/
        返回 ticket
        """
        url = f"{self.host}/space/api/export/create/"
        body = {
            "token": excel_token,
            "type": "sheet",
            "file_extension": "xlsx",
            "event_source": "1",
        }
        log.info(f"📤 创建 Excel 导出任务: token={excel_token}")
        resp = self.session.post(url, json=body)
        resp.raise_for_status()
        result = resp.json()
        ticket = result.get("data", {}).get("ticket")
        if not ticket:
            raise RuntimeError(f"创建导出任务失败: {resp.text[:300]}")
        log.info(f"   ✅ ticket={ticket}")
        return ticket

    # ----- Step 7: 轮询导出结果 -----
    def poll_export_result(self, ticket: str, timeout: int = 600) -> tuple[str | None, str]:
        """
        GET /space/api/export/result/{ticket}
        返回 (file_download_token, job_error_msg)
        """
        url = f"{self.host}/space/api/export/result/{ticket}"
        params = {"token": self.excel_token, "type": "sheet"}

        start = time.time()
        while time.time() - start < timeout:
            resp = self.session.get(url, params=params)
            resp.raise_for_status()
            result = resp.json()
            data = result.get("data", {})
            job_error_msg = data.get("job_error_msg", "")
            file_download_token = data.get("file_download_token")

            log.info(f"   ⏳ 导出状态: job_error_msg='{job_error_msg}'")
            if job_error_msg == "success" and file_download_token:
                log.info(f"   ✅ 导出完成: file_download_token={file_download_token}")
                return file_download_token, job_error_msg
            if job_error_msg and job_error_msg != "success":
                log.error(f"   ❌ 导出错误: {job_error_msg}")
                return None, job_error_msg

            time.sleep(CONFIG["poll_interval"])

        raise TimeoutError(f"导出超时 ({timeout}s)")

    # ----- Step 8: 下载 Excel 并解析金标准 -----
    def download_and_parse_excel(
        self, file_download_token: str, sheet_name: str, file_name: str
    ) -> tuple[dict, int]:
        """
        下载导出的 Excel 文件并解析金标准数据。

        原 Apifox 调用 analysis_excel.py 实现，此处内联实现。

        返回 (gold_standard_dict, file_count)
        """
        # 飞书文件下载 URL
        download_url = f"{self.host}/space/api/export/file/download/"
        params = {"file_download_token": file_download_token}

        log.info(f"📥 下载金标准 Excel...")
        resp = self.session.get(download_url, params=params)
        resp.raise_for_status()

        # 保存到临时文件
        download_dir = Path(self.download_path) if self.download_path else Path(tempfile.gettempdir()) / "feishu_download"
        download_dir.mkdir(parents=True, exist_ok=True)
        excel_path = download_dir / file_name
        with open(excel_path, "wb") as f:
            f.write(resp.content)
        log.info(f"   Excel 已保存: {excel_path} ({len(resp.content)} bytes)")

        # 解析
        gold_standard, file_count = read_excel_gold_standard(str(excel_path), sheet_name)
        return gold_standard, file_count


# ============================================================
# Phase 2: AZ 审核客户端
# ============================================================
class AZClient:
    """AZ 审核系统 API 客户端"""

    def __init__(self, config: dict):
        self.base_url = config["az_base_url"].rstrip("/")
        self.verify_ssl = config["verify_ssl"]
        self.audit_defaults = config["audit_defaults"]

        self.session = requests.Session()
        self.session.verify = self.verify_ssl
        self.session.headers.update(
            {
                "User-Agent": "AZ-Audit-Test/1.0 (Python)",
                "Content-Type": "application/json",
            }
        )
        self.token: str | None = None

    # ----- Step 9: 登录 -----
    def login(self, email: str, password: str) -> str:
        url = f"{self.base_url}/auth/email/login"
        body = {"email": email, "password": password}
        log.info(f"🔑 登录中... (email={email})")
        resp = self.session.post(url, json=body)
        resp.raise_for_status()
        data = resp.json()
        token = data.get("data", {}).get("token")
        if not token:
            raise RuntimeError(f"登录失败: {resp.text[:300]}")
        self.token = f"Bearer {token}"
        log.info("✅ 登录成功")
        return self.token

    # ----- Step 11: 上传文件 -----
    def upload_file(self, file_path: str) -> str | None:
        url = f"{self.base_url}/oss/upload"
        headers = {"Authorization": self.token} if self.token else {}
        file_name = Path(file_path).name
        log.info(f"📤 上传: {file_name}")

        try:
            with open(file_path, "rb") as f:
                resp = self.session.post(
                    url, headers=headers, files={"file": (file_name, f)}
                )
            resp.raise_for_status()
            file_id = resp.json().get("data")
            if file_id:
                log.info(f"   ✅ file_id={file_id}")
                return file_id
            log.warning(f"   ⚠️ 上传返回空: {resp.text[:200]}")
            return None
        except requests.RequestException as e:
            log.error(f"   ❌ 上传失败: {e}")
            return None

    # ----- Step 11: 创建审核任务 -----
    def create_audit_task(self, file_id: str, file_name: str) -> str | None:
        url = f"{self.base_url}/audit/management/add"
        headers = {
            "Authorization": self.token,
            "Content-Type": "application/json",
        } if self.token else {"Content-Type": "application/json"}

        body = {
            "file_id": file_id,
            "file_name": file_name,
            **self.audit_defaults,
        }
        log.info(f"📝 创建审核任务: {file_name}")

        try:
            resp = self.session.post(url, json=body, headers=headers)
            resp.raise_for_status()
            task_id = resp.json().get("data")
            if task_id:
                log.info(f"   ✅ task_id={task_id}")
                return task_id
            log.warning(f"   ⚠️ 返回空: {resp.text[:200]}")
            return None
        except requests.RequestException as e:
            log.error(f"   ❌ 创建失败: {e}")
            return None

    # ----- Step 12: 获取审核结果 -----
    def get_audit_detail(self, task_id: str) -> dict | None:
        url = f"{self.base_url}/audit/management/detail"
        headers = {
            "Authorization": self.token,
            "Content-Type": "application/json",
        } if self.token else {"Content-Type": "application/json"}
        body = {"id": task_id}

        try:
            resp = self.session.post(url, json=body, headers=headers)
            resp.raise_for_status()
            return resp.json()
        except requests.RequestException as e:
            log.error(f"   ❌ 获取审核详情失败 (id={task_id}): {e}")
            return None


# ============================================================
# 工具函数
# ============================================================
def read_excel_gold_standard(
    filepath: str, sheet_name: str
) -> tuple[dict[str, list[dict]], int]:
    """
    读取金标准 Excel，返回 (merged_result, file_count)。

    merged_result: {"文件名": [{"审核点": ..., "期望值": ..., ...}, ...], ...}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}")

    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            row_dict = {headers[i]: value for i, value in enumerate(row) if i < len(headers) and headers[i] is not None}
            rows.append(row_dict)
    wb.close()

    merged_result: dict[str, list[dict]] = {}
    for row in rows:
        file_name = row.get("文件名") or row.get("文件名称") or row.get("file_name")
        if file_name is None:
            continue
        merged_result.setdefault(file_name, []).append(row)

    log.info(f"✅ 金标准: {len(rows)} 条记录, {len(merged_result)} 个文件")
    return merged_result, len(merged_result)


def scan_files(directory: str) -> list[str]:
    dir_path = Path(directory)
    if not dir_path.exists():
        raise FileNotFoundError(f"目录不存在: {directory}")
    files = [
        str(p.resolve())
        for p in dir_path.iterdir()
        if p.is_file() and p.name != ".DS_Store"
    ]
    log.info(f"📂 扫描到 {len(files)} 个文件")
    return files


def compare_with_gold_standard(
    client: AZClient,
    task_queue: list[dict],
    gold_standard: dict[str, list[dict]],
) -> list[dict]:
    """
    获取审核结果并与金标准对比。
    返回对比结果列表。
    """
    results = []
    for task in task_queue:
        task_id = task["id"]
        file_name = task["fileName"]
        log.info(f"🔍 对比: {file_name}")

        detail = client.get_audit_detail(task_id)
        if detail is None:
            results.append({"file_name": file_name, "task_id": task_id, "status": "error", "error": "无法获取审核结果"})
            continue

        audit_data = detail.get("data", {})
        gold_rows = gold_standard.get(file_name) or gold_standard.get(Path(file_name).stem, [])

        if not gold_rows:
            log.warning(f"   ⚠️ 无金标准数据")
            results.append({"file_name": file_name, "task_id": task_id, "status": "no_gold_standard", "audit_data": audit_data})
            continue

        row_results = []
        for gold_row in gold_rows:
            point = gold_row.get("审核点", gold_row.get("审核项", ""))
            expected = gold_row.get("期望值", gold_row.get("期望结果", gold_row.get("金标准", "")))
            actual = _extract_audit_value(audit_data, gold_row)
            passed = str(actual) == str(expected) if actual is not None else False
            row_results.append({"审核点": point, "期望值": expected, "实际值": actual, "通过": passed})

        passed = sum(1 for r in row_results if r["通过"])
        total = len(row_results)
        log.info(f"   {'✅' if passed == total else '❌'} {passed}/{total}")
        results.append({
            "file_name": file_name, "task_id": task_id, "status": "completed",
            "points": row_results, "passed": passed, "total": total, "pass_rate": f"{passed}/{total}",
        })
    return results


def _extract_audit_value(audit_data: dict, gold_row: dict) -> Any | None:
    field_path = gold_row.get("字段名") or gold_row.get("field") or gold_row.get("审核字段")
    if field_path:
        value = audit_data
        for key in str(field_path).split("."):
            if isinstance(value, dict):
                value = value.get(key)
            else:
                return None
        return value
    # 递归搜索匹配的 key
    return _deep_search(audit_data, gold_row.get("审核点", ""))


def _deep_search(data: dict, target: str) -> Any | None:
    if not isinstance(data, dict):
        return None
    for key, value in data.items():
        if target in str(key):
            return value
        if isinstance(value, dict):
            result = _deep_search(value, target)
            if result is not None:
                return result
    return None


def generate_report(results: list[dict], output_path: str = "audit_comparison_report.json"):
    total = len(results)
    completed = sum(1 for r in results if r["status"] == "completed")
    errors = sum(1 for r in results if r["status"] == "error")
    no_gold = sum(1 for r in results if r["status"] == "no_gold_standard")
    all_passed = sum(r.get("passed", 0) for r in results if r["status"] == "completed")
    all_total = sum(r.get("total", 0) for r in results if r["status"] == "completed")

    summary = {
        "总文件数": total,
        "成功对比": completed,
        "获取失败": errors,
        "无金标准": no_gold,
        "审核点通过率": f"{all_passed}/{all_total}" if all_total else "N/A",
    }
    report = {"summary": summary, "details": results}

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    log.info("=" * 60)
    log.info("📊 对比报告摘要")
    for k, v in summary.items():
        log.info(f"  {k}: {v}")
    log.info(f"  详细结果: {output_path}")
    return report


# ============================================================
# 主流程
# ============================================================
def main():
    cfg = CONFIG

    # ---- Phase 1: 从飞书获取金标准 Excel ----
    log.info("=" * 60)
    log.info("Phase 1: 从飞书下载金标准 Excel")
    log.info("=" * 60)

    fs = FeishuClient(cfg)

    # Step 1-2: 创建文件夹下载 & 等待
    invoke_code = fs.create_folder_download()
    fs.poll_folder_download(invoke_code)

    # Step 3-4: 打包 zip & 等待
    zip_code = fs.create_zip()
    fs.poll_zip_status(zip_code)

    # Step 5: 下载 zip，解压获取 Excel token
    excel_token = fs.download_zip_and_get_excel_token(zip_code)
    # 如果步骤 5 返回了新的 token，用它
    cfg["feishu_excel_token"] = excel_token
    fs.excel_token = excel_token

    # Step 6-7: 导出 Excel & 等待
    ticket = fs.create_export(excel_token)
    file_download_token, job_error_msg = fs.poll_export_result(ticket)

    if job_error_msg != "success" or not file_download_token:
        log.error("❌ Excel 导出失败，无法继续")
        sys.exit(1)

    # Step 8: 下载并解析金标准 Excel
    gold_standard, gold_file_count = fs.download_and_parse_excel(
        file_download_token, cfg["gold_standard_sheetname"], cfg["gold_standard_filename"]
    )

    # ---- Phase 2: AZ 审核流程 ----
    log.info("=" * 60)
    log.info("Phase 2: AZ 审核流程")
    log.info("=" * 60)

    if not cfg["az_email"] or not cfg["az_password"]:
        log.error("❌ 请设置 AZ_EMAIL 和 AZ_PASSWORD 环境变量")
        sys.exit(1)

    az = AZClient(cfg)

    # Step 9: 登录
    az.login(cfg["az_email"], cfg["az_password"])

    # Step 10: 扫描文件
    file_paths = scan_files(cfg["file_dir"])
    if not file_paths:
        log.error("❌ 没有待审核文件")
        sys.exit(1)

    # Step 11: 遍历 — 上传 + 创建审核任务
    log.info("--- 上传文件 & 创建审核任务 ---")
    task_queue: list[dict] = []
    for fp in file_paths:
        fname = Path(fp).name
        fid = az.upload_file(fp)
        if not fid:
            continue
        tid = az.create_audit_task(fid, fname)
        if tid:
            task_queue.append({"id": tid, "fileName": fname})

    log.info(f"📋 共创建 {len(task_queue)} 个审核任务")

    # 等待审核完成
    log.info("--- 等待审核完成 ---")
    pending = {t["id"] for t in task_queue}
    for attempt in range(cfg["max_poll_attempts"]):
        if not pending:
            break
        still_pending = set()
        for tid in pending:
            detail = az.get_audit_detail(tid)
            if detail is None:
                still_pending.add(tid)
                continue
            status = (detail.get("data", {}).get("status")
                      or detail.get("data", {}).get("audit_status")
                      or detail.get("data", {}).get("state", ""))
            if status not in ("completed", "done", "finished", "审核完成", "已完成"):
                still_pending.add(tid)
        pending = still_pending
        if pending:
            log.info(f"   ⏳ {len(pending)} 个未完成 (第 {attempt+1}/{cfg['max_poll_attempts']} 次)")
            time.sleep(cfg["poll_interval"])

    # Step 12: 对比
    log.info("--- 获取审核结果 & 对比金标准 ---")
    results = compare_with_gold_standard(az, task_queue, gold_standard)

    # 报告
    generate_report(results)

    all_pass = all(
        r["status"] == "completed" and r.get("passed", 0) == r.get("total", 0)
        for r in results
    )
    sys.exit(0 if all_pass else 1)


if __name__ == "__main__":
    main()
